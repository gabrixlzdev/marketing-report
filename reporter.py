from openpyxl import Workbook
from tkinter import filedialog
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

def generate_report(summary: dict, df_filtered: dict, column_campaings: list, selected_campaings: list, config: dict) -> None:
    # This function will generate the report based on the summary of the data
    # You can implement the logic to create a new sheet in the workbook and write the summary of the data to it as needed

    wb= Workbook() # create a new workbook (sheet) to store the summary of the data
    ws = wb.active # get the active sheet in the workbook

    output_path = filedialog.asksaveasfilename(
        title= "Save report as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    labels = {
        "total_spend": "Total Spend (R$)",
        "total_impressions": "Impressions",
        "total_reach": "Reach",
        "total_clicks": "Clicks",
        "total_conversions": "Conversions",
        "avg_cpl": "Avg CPL (R$)",
        "real_cpl": "Real CPL (R$)",
        "total_new_contacts": "New Contacts"
    }

    alfabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ" # create a string of the alphabet to use for the column names in the sheet

    ws.title = "Summary" # set the title of the sheet to "Summary"
    font="Times New Roman" # set the font of the sheet to "Times New Roman"

    i=0
    last_column = alfabet[len(summary)-1] # get the last column of the sheet based on the number of keys in the summary dictionary
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin) # create a border style for the cells in the sheet
  
    keys = list(labels.values()) 
    values = list(summary.values()) #get the values of the summary dictionary as a list

    ws. merge_cells(f"A1:{last_column}1") # merge the cells in the first row of the sheet to create a title for the summary
    ws["A1"].value = "Summary of the data" # set the value of the merged cells to "Summary of the data"
    ws["A1"].font = Font(name=font ,size=14, bold=True, color = "FFFFFF") # set the font of the title to size 14 and bold
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center") # set the alignment of the title to center
    ws["A1"].fill = PatternFill("solid", start_color = "1F4E79") # set the background color of the title to a dark blue
    ws["A1"].border = border # set the border of the title to the thin border style
    ws.row_dimensions[1].height = 30 # set the height of the first row to 30

    while i < len(summary):

        col_letter = get_column_letter(i+1) # get the column letter based on the index of the loop
        col_width = max(len(str(keys[i])), len(str(values[i]))) + 2 # calculate the width of the column based on the length of the key and value in the summary dictionary
        ws.column_dimensions[col_letter].width = col_width # set the width of the column in the sheet

        ws.cell(row=2, column=i+1).value = keys[i] #write the keys of the summary dictionary to the first row of the sheet

        ws.cell(row=3, column=i+1).value = values[i] #write the values of the summary dictionary to the second row of the sheet
        
        ws.cell(row=2, column=i+1).font = Font(name=font,bold=True, color="FFFFFF")
        ws.cell(row=2, column=i+1).fill = PatternFill("solid", start_color="1F4E79")
        ws.cell(row=2, column=i+1).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=i+1).border = border
        
        ws.cell(row=3, column=i+1).fill = PatternFill("solid", start_color="D6E4F0")
        ws.cell(row=3, column=i+1).alignment = Alignment(horizontal="center")
        ws.cell(row=3, column=i+1).font = Font(name=font)
        ws.cell(row=3, column=i+1).border = border
        ws.cell (row=3, column=i+1).number_format = "#,##0.00" if isinstance(values[i], float) else "#,##0" # set the number format of the cell to have two decimal places if the value is a float, or no decimal places if the value is an integer

        i+=1
    
    for campaign in selected_campaings:
        ws_campaign = wb.create_sheet(title=campaign) # create a new sheet for each selected campaign in the configuration
        df_campaign = df_filtered[df_filtered[column_campaings] == campaign] # filter the data frame to get the data for the current campaign
        
        #df_filtered -> eg: ["T CBO Mensagem", "T CBO Mensagem", "CBO T"]
        #df_filtered[column_campaings] == campaign -> eg: ["T CBO Mensagem", "T CBO Mensagem", "CBO T"] == "T CBO Mensagem" -> [True, True, False]
        # You can implement the logic to write the data for each campaign to the corresponding sheet as needed
        # df_filtered will retorn a data frame with the data for the current campaign when the return of the filter is True

        i=0
  
        keys = df_campaign.columns.tolist() #get the columns of the data frame as a list
        values = list(df_campaign.values.tolist()[0]) #get the values of the summary dictionary as a list
        last_column_campaign = get_column_letter(len(df_campaign.columns)) # get the last column of the sheet based on the number of columns in the data frame for the current campaign
        ws_campaign.merge_cells(f"A1:{last_column_campaign}1") # merge the cells in the first row of the sheet to create a title for the campaign data

    
        ws_campaign["A1"].value = campaign # set the value of the merged cells to the campaign name
        ws_campaign["A1"].font = Font(name=font ,size=14, bold=True, color = "FFFFFF") # set the font of the title to size 14 and bold
        ws_campaign["A1"].alignment = Alignment(horizontal="center", vertical="center") # set the alignment of the title to center
        ws_campaign["A1"].fill = PatternFill("solid", start_color = "1F4E79") # set the background color of the title to a dark blue
        ws_campaign["A1"].border = border # set the border of the title to the thin border style
        ws_campaign.row_dimensions[1].height = 30 # set the height of the first row to 30

        while i < len(df_campaign.columns):

            col_letter = get_column_letter(i+1) # get the column letter based on the index of the loop
            col_width = max(len(str(keys[i])), len(str(values[i]))) + 2 # calculate the width of the column based on the length of the key and value in the summary dictionary
            ws_campaign.column_dimensions[col_letter].width = col_width # set the width of the column in the sheet

            ws_campaign.cell(row=2, column=i+1).value = keys[i] #write the keys of the summary dictionary to the first row of the sheet

            ws_campaign.cell(row=3, column=i+1).value = values[i] #write the values of the summary dictionary to the second row of the sheet
        
            ws_campaign.cell(row=2, column=i+1).font = Font(name=font,bold=True, color="FFFFFF")
            ws_campaign.cell(row=2, column=i+1).fill = PatternFill("solid", start_color="1F4E79")
            ws_campaign.cell(row=2, column=i+1).alignment = Alignment(horizontal="center")
            ws_campaign.cell(row=2, column=i+1).border = border
        
            ws_campaign.cell(row=3, column=i+1).fill = PatternFill("solid", start_color="D6E4F0")
            ws_campaign.cell(row=3, column=i+1).alignment = Alignment(horizontal="center")
            ws_campaign.cell(row=3, column=i+1).font = Font(name=font)
            ws_campaign.cell(row=3, column=i+1).border = border
            ws_campaign.cell (row=3, column=i+1).number_format = "#,##0.00" if isinstance(values[i], float) else "#,##0" # set the number format of the cell to have two decimal places if the value is a float, or no decimal places if the value is an integer

            i+=1
    
    headers= ["Campaign Name","Spend", "conversions", "CPL"] # create a list of headers for the campaign data
    for i in range(len(headers)): #range(len(headers)) -> range(4) -> [0, 1, 2, 3]
        ws.cell(row=6, column=i+1).value = headers[i] # write the headers to the fourth row of the sheet
        ws.cell(row=6, column=i+1).font = Font(name=font ,size=14, bold=True, color = "FFFFFF") # set the font of the title to size 14 and bold
        ws.cell(row=6, column=i+1).fill = PatternFill("solid", start_color = "1F4E79") # set the background color of the title to a dark blue
        ws.cell(row=6, column=i+1).alignment = Alignment(horizontal="center")
        ws.cell(row=6, column=i+1).border = border

    i=0
    while i< len(selected_campaings):

        campaign = selected_campaings[i] # get the current campaign name from the list of selected campaigns
        df_campaign = df_filtered[df_filtered[column_campaings] == campaign] # filter the data frame to get the data for the current campaign
        
        ws.cell(row=i+7, column=1).value = campaign # write the campaign name to the first column of the sheet
        ws.cell(row=i+7, column=2).value = df_campaign[config["spend"]].values[0] # write the total spend for the current campaign to the second column of the sheet
        ws.cell(row=i+7, column=3).value = df_campaign[config["conversions"]].values[0] # write the total conversions for the current campaign to the third column of the sheet
        ws.cell(row=i+7, column=4).value = df_campaign[config["cost_per_lead"]].values[0] # write the cost per lead for the current campaign to the fourth column of the sheet

        j=1
        while j<=4:
            ws.cell(row=i+7, column=j).font = Font(name=font) # set the font of the cells to the specified font
            ws.cell(row=i+7, column=j).fill = PatternFill("solid", start_color="D6E4F0")
            ws.cell(row=i+7, column=j).alignment = Alignment(horizontal="center")
            ws.cell(row=i+7, column=j).border = border

            
            j+=1

        i+=1

    #graphics to visualize the data for each campaign
    last_row = len(selected_campaings) + 6 # calculate the last row of the campaign data in the sheet

    data= Reference(ws, min_col=2, min_row=6, max_row=last_row) # create a reference to the campaign data in the sheet
    x_axis= Reference(ws, min_col=1, min_row=7, max_row=last_row) # create a reference to the campaign names in the sheet

    #bar chart to visualize the total spend for each campaign

    chart_bar = BarChart() # create a bar chart to visualize the campaign data
    chart_bar.add_data(data, titles_from_data=True) # add the campaign data to the chart
    chart_bar.set_categories(x_axis) # set the x-axis of the chart to the campaign names
    chart_bar.title = "Value spend by campaign" # set the title of the chart
    ws.add_chart(chart_bar, "F6") # add the chart to the sheet at the specified position

    #pie chart to visualize the total conversions for each campaign

    chart_pie=PieChart() # create a pie chart to visualize the campaign data
    chart_pie.add_data(data, titles_from_data=True) # add the campaign data to the chart
    chart_pie.set_categories(x_axis) # set the x-axis of the chart to the campaign names
    chart_pie.title = "Budget allocation by campaign" # set the title of the chart
    ws.add_chart(chart_pie, "M20") # add the chart to the sheet at the specified position

    wb.save(output_path) # save the workbook to the specified path
    



